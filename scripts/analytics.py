# scripts/analytics.py
import os
import argparse
import pandas as pd
import matplotlib.pyplot as plt
from sqlalchemy import create_engine, text
from datetime import datetime

# ========== НАСТРОЙКИ ==========
DB_USER = "postgres"
DB_PASS = "kazbekova2005D." 
DB_HOST = "localhost"
DB_PORT = "5432"
DB_NAME = "ecommerce_olist"

SQLALCHEMY_URI = f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

CHARTS_DIR = "charts"
EXPORTS_DIR = "exports"

plt.rcParams.update({
    "figure.figsize": (10, 6),
    "axes.grid": True,
    "axes.titlesize": 14,
    "axes.labelsize": 12
})


# ========== ВСПОМОГАТЕЛЬНЫЕ ==========
def ensure_dirs():
    os.makedirs(CHARTS_DIR, exist_ok=True)
    os.makedirs(EXPORTS_DIR, exist_ok=True)

def get_engine():
    return create_engine(SQLALCHEMY_URI)

def q(sql: str, params=None) -> pd.DataFrame:
    with get_engine().connect() as conn:
        return pd.read_sql(text(sql), conn, params=params)

def save_chart(fig, filename: str, rows_count: int, description: str):
    path = os.path.join(CHARTS_DIR, filename)
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    print(f"[OK] {filename}: {rows_count} rows — {description}")


# ========== 6 ГРАФИКОВ ==========
def pie_payments_by_type():
    sql = """
    SELECT op.payment_type, COUNT(*) AS n
    FROM order_payments op
    JOIN orders o   ON o.order_id = op.order_id
    JOIN customers c ON c.customer_id = o.customer_id
    GROUP BY op.payment_type
    ORDER BY n DESC;
    """
    df = q(sql)
    fig, ax = plt.subplots()
    ax.pie(df["n"], labels=df["payment_type"], autopct="%1.1f%%", startangle=90)
    ax.set_title("Как платят покупатели: распределение способов оплаты")
    save_chart(fig, "01_pie_payments.png", len(df), "Pie chart: способы оплаты")


def bar_top_categories_revenue():
    sql = """
    SELECT p.product_category_name,
           SUM(oi.price + oi.freight_value) AS revenue
    FROM order_items oi
    JOIN products p ON p.product_id = oi.product_id
    JOIN orders o   ON o.order_id = oi.order_id
    GROUP BY p.product_category_name
    ORDER BY revenue DESC
    LIMIT 10;
    """
    df = q(sql)
    fig, ax = plt.subplots()
    ax.bar(df["product_category_name"], df["revenue"])
    ax.set_title("ТОП-10 товарных категорий по выручке")
    ax.set_xlabel("Категория")
    ax.set_ylabel("Выручка")
    plt.xticks(rotation=45, ha="right")
    save_chart(fig, "02_bar_categories.png", len(df), "Bar chart: категории по выручке")


def barh_city_revenue():
    sql = """
    SELECT c.customer_city,
           SUM(oi.price + oi.freight_value) AS revenue
    FROM customers c
    JOIN orders o      ON o.customer_id = c.customer_id
    JOIN order_items oi ON oi.order_id = o.order_id
    GROUP BY c.customer_city
    ORDER BY revenue DESC
    LIMIT 10;
    """
    df = q(sql)
    fig, ax = plt.subplots()
    ax.barh(df["customer_city"], df["revenue"])
    ax.set_title("ТОП-10 городов с наибольшими продажами")
    ax.set_xlabel("Выручка")
    ax.set_ylabel("Город")
    ax.invert_yaxis()
    save_chart(fig, "03_barh_cities.png", len(df), "Horizontal bar: города по выручке")


def line_monthly_orders_and_revenue():
    sql = """
    SELECT
      to_char(date_trunc('month', o.order_purchase_timestamp), 'YYYY-MM') AS order_month,
      COUNT(DISTINCT o.order_id) AS orders_cnt,
      SUM(oi.price + oi.freight_value) AS revenue
    FROM orders o
    JOIN order_items oi   ON oi.order_id = o.order_id
    JOIN order_payments op ON op.order_id = o.order_id
    GROUP BY order_month
    ORDER BY order_month;
    """
    df = q(sql)
    fig, ax = plt.subplots()
    ax.plot(df["order_month"], df["orders_cnt"], label="Заказы")
    ax.plot(df["order_month"], df["revenue"], label="Выручка")
    ax.set_title("Как менялись заказы и выручка по месяцам")
    ax.set_xlabel("Месяц")
    ax.set_ylabel("Значение")
    plt.xticks(rotation=45, ha="right")
    ax.legend()
    save_chart(fig, "04_line_orders.png", len(df), "Line chart: заказы и выручка по месяцам")


def hist_order_totals():
    sql = """
    WITH order_totals AS (
      SELECT o.order_id,
             SUM(oi.price + oi.freight_value) AS order_total
      FROM orders o
      JOIN order_items oi ON oi.order_id = o.order_id
      GROUP BY o.order_id
    )
    SELECT * FROM order_totals;
    """
    df = q(sql)
    fig, ax = plt.subplots()
    ax.hist(df["order_total"].astype(float), bins=30)
    ax.set_title("Как распределены суммы заказов (чек)")
    ax.set_xlabel("Сумма заказа")
    ax.set_ylabel("Частота")
    save_chart(fig, "05_hist_order_totals.png", len(df), "Histogram: распределение сумм заказов")


def scatter_price_vs_freight():
    sql = """
    SELECT oi.price, oi.freight_value
    FROM order_items oi
    JOIN products p ON p.product_id = oi.product_id
    JOIN orders o   ON o.order_id = oi.order_id
    WHERE oi.price IS NOT NULL AND oi.freight_value IS NOT NULL
    LIMIT 5000;
    """
    df = q(sql)
    fig, ax = plt.subplots()
    ax.scatter(df["price"].astype(float), df["freight_value"].astype(float), alpha=0.5)
    ax.set_title("Связь: цена товара и стоимость доставки (сэмпл)")
    ax.set_xlabel("Цена товара")
    ax.set_ylabel("Стоимость доставки")
    save_chart(fig, "06_scatter_price_freight.png", len(df), "Scatter: price vs freight")


# ========== PLOTLY SLIDER ==========
def show_plotly_time_slider():
    import plotly.express as px
    sql = """
    SELECT
      to_char(date_trunc('month', o.order_purchase_timestamp), 'YYYY-MM') AS order_month,
      p.product_category_name,
      SUM(oi.price + oi.freight_value) AS revenue
    FROM orders o
    JOIN order_items oi ON oi.order_id = o.order_id
    JOIN products p     ON p.product_id = oi.product_id
    GROUP BY order_month, p.product_category_name;
    """
    df = q(sql)
    top5 = (df.groupby("product_category_name")["revenue"]
              .sum()
              .sort_values(ascending=False)
              .head(5)
              .index.tolist())
    dff = df[df["product_category_name"].isin(top5)].copy()
    fig = px.bar(
        dff.sort_values(["order_month", "revenue"], ascending=[True, False]),
        x="product_category_name", y="revenue",
        animation_frame="order_month",
        title="Выручка ТОП-5 категорий по месяцам (ползунок времени)",
        labels={"product_category_name": "Категория", "revenue": "Выручка"}
    )
    fig.show()


# ========== ЭКСПОРТ В EXCEL ==========
def export_to_excel(dataframes: dict, filename: str):
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import ColorScaleRule

    path = os.path.join(EXPORTS_DIR, filename)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            rng = f"{col[0].column_letter}2:{col[0].column_letter}{ws.max_row}"
            rule = ColorScaleRule(start_type="min", start_color="FFAA0000",
                                  mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                                  end_type="max", end_color="FF00AA00")
            ws.conditional_formatting.add(rng, rule)
    wb.save(path)
    print(f"[OK] Создан файл {filename}")


# ========== ДЕМО INSERT ==========
def demo_insert_and_regenerate():
    with get_engine().begin() as conn:
        order_id = conn.execute(text("SELECT order_id FROM orders LIMIT 1")).scalar()
        conn.execute(text("""
            INSERT INTO order_payments(order_id, payment_type, payment_installments, payment_value)
            VALUES (:oid, 'credit_card', 1, 123.45)
        """), {"oid": order_id})
        print("[OK] Добавлена 1 строка в order_payments")
    pie_payments_by_type()
    print("[OK] Пересчитан график pie")


# ========== MAIN ==========
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--show-slider", action="store_true")
    parser.add_argument("--demo-update", action="store_true")
    args = parser.parse_args()

    ensure_dirs()

    if args.show_slider:
        show_plotly_time_slider()
        return
    if args.demo_update:
        demo_insert_and_regenerate()
        return

    pie_payments_by_type()
    bar_top_categories_revenue()
    barh_city_revenue()
    line_monthly_orders_and_revenue()
    hist_order_totals()
    scatter_price_vs_freight()

    dfs = {
        "payments": q("SELECT payment_type, COUNT(*) AS n FROM order_payments GROUP BY payment_type"),
        "categories": q("SELECT p.product_category_name, SUM(oi.price+oi.freight_value) AS revenue "
                        "FROM order_items oi JOIN products p ON p.product_id=oi.product_id "
                        "GROUP BY p.product_category_name ORDER BY revenue DESC LIMIT 10"),
        "monthly_orders": q("SELECT to_char(date_trunc('month', order_purchase_timestamp), 'YYYY-MM') AS m, "
                            "COUNT(*) FROM orders GROUP BY m ORDER BY m")
    }
    export_to_excel(dfs, f"ecommerce_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


if __name__ == "__main__":
    main()
